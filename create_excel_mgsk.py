from pdf2docx import Converter
import logging
import re
import os
from collections import defaultdict
from materials_from_table import extract_materials
from act_from_table import reformat_table_11_17, format_table
from merge_tables import merge_t
from technology_parser import get_MGSK
import openpyxl
from act_from_text import get_info_from_text
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border, Color, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from itertools import product
import string
from itertools import chain
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

FIELDS_TECH = ['Скважина', 'Приемистость скважины на 1-й скорости', 'Приемистость скважины на 2-й скорости',
               'Приемистость скважины на 3-й скорости', 'Общий объем закачки МГС-К', 'Количество стадий закачки',
               'Объем закачки на первой стадии', 'Объем закачки на второй стадии', 'концентрация ПАА на первой стадии',
               'концентрация ПАА на второй стадии', 'концентрация ПАВ на первой стадии',
               'концентрация ПАВ на второй стадии', 'концентрация СА (СКА) на первой стадии',
               'концентрация СА (СКА) на второй стадии', 'концентрация ХКК на первой стадии',
               'концентрация ХКК на второй стадии', 'Давление закачки', 'Объем продавки после первой фазы',
               'Объем финальной продавки', 'Приемистость скважины на 1-й скорости после закачки',
               'Приемистость скважины на 2-й скорости после закачки',
               'Приемистость скважины на 3-й скорости после закачки']
for i in range(1, 10):
    FIELDS_TECH += [f'Материал {i}', f'Плотность материала {i}', f'Количество материала {i}']
FIELDS_TECH += ['Согласовано', 'Комментарий']


def replace_empty_to_not_find(data):
    for i in range(len(data)):
        if len(str(data[i])) == 0:
            data[i] = 'н/д'
    return data


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_width_custom(ws):
    columns_40 = ['B', 'C', 'D', 'T', 'U', 'V']
    first_letters = ['', 'A']
    second_letters = [list(string.ascii_uppercase), list(string.ascii_uppercase)[:-1]]
    columns_15 = []
    for i, first in enumerate(first_letters):
        for second in second_letters[i]:
            columns_15.append(first + second)
    for letter in columns_40:
        columns_15.remove(letter)
    for col in columns_40:
        ws.column_dimensions[col].width = 40
    for col in columns_15:
        ws.column_dimensions[col].width = 15


def format_data(value):
    value = str(value)
    if re.fullmatch(r'\d+', value):
        return int(value)
    elif re.fullmatch(r'\d+\.\d+', value):
        return float(value)
    elif re.fullmatch(r'\d+,\d+', value):
        value = value.replace(',', '.')
        return float(value)
    return value


# Character range function


# Example run

def format_percentage(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            value = str(cell.value)
            if re.fullmatch(r'\d+[,.]?\d+%', value):
                cell.value = value.replace(',', '.')
                cell.number_format = FORMAT_PERCENTAGE
                #cell.value = value.replace('.', ',')


def set_hor_center(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


def create_xlsx_mgsk(data_samples):
    data_samples = dict(zip(data_samples.keys(), list(map(lambda x: list(map(format_data, x)), data_samples.values()))))
    # создаю книгу
    book = openpyxl.Workbook()

    # по умолчанию создается с таблицей Sheet
    # print(book.sheetnames)
    book.remove(book.active)
    # создаю таблицы
    sheet = book.create_sheet('Лист1')
    data_keys = list(data_samples.keys())
    print(data_keys)
    sheet.append(data_keys)
    for i in range(len(data_samples['Скважина'])):  # получаю данные
        # book.guess_types = True
        print(data_samples.values())
        try:
            values = [d[i] for d in list(data_samples.values())]
        except IndexError:
            print('IndexError in ', i)
            print(data_samples['Скважина'][i])
        # append the `generator values`
        sheet.append(values)
        # записываю данные в строки таблицы
    set_width_custom(sheet)
    full_table_range = f"A1:AY{len(data_samples['Скважина']) + 1}"
    set_border(sheet, full_table_range)
    # центрирование
    set_hor_center(sheet, full_table_range)
    format_percentage(sheet, full_table_range)
    path_result = f'МГСК/МГСК.xlsx'
    os.makedirs(os.path.dirname(path_result), exist_ok=True)
    book.save(path_result)


if __name__ == "__main__":
    with open('path_directory.txt', encoding='utf-8', mode='r') as f:
        directory = f.readline()
    data = defaultdict(list)
    for field, idx in zip(FIELDS_TECH, list(range(1, len(FIELDS_TECH) + 1))):
        data[field].append(idx)
    print(data)
    break_idx = 0
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                path_file = os.path.join(root, file)
                data = get_MGSK(data, path_file)
                """
                материалы из таблицы
                """
                materials = extract_materials(path_file)

                for i, material in enumerate(materials):
                    idx = i + 1
                    material_new = replace_empty_to_not_find(material)
                    data[f'Материал {idx}'].append(material_new[0])
                    data[f'Плотность материала {idx}'].append(material_new[1])
                    data[f'Количество материала {idx}'].append(material_new[2])
                for i, k in enumerate(data.keys()):
                    if len(data[k]) < len(data['Скважина']):
                        data[k].append('н/д')
                # break_idx += 1
                # if break_idx > 20:
                #     break
    create_xlsx_mgsk(data)
