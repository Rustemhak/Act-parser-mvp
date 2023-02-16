import os
import openpyxl
from act_from_text import get_info_from_text
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border, Color, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_color(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(color="00FF0000")


def merge_2_row(ws, start_column, end_column, row_num):
    for row_column in range(start_column, end_column + 1):
        ws.merge_cells(start_row=row_num, start_column=row_column, end_row=row_num + 1, end_column=row_column)


def set_width(ws):
    columns_30 = "DEFGHIJK"
    column_20 = "L"
    for col in columns_30:
        ws.column_dimensions[col].width = 30
    ws.column_dimensions[column_20].width = 20


def create_xlsx(path, data_samples, table2, table3):
    """
    Создание xlsx файла и запись в него
       Редактирую таблицу
    - двигаю данные
    - добавляю колонки и строки
    - записываю данные в ячейки
    https://openpyxl.readthedocs.io/en/stable/editing_worksheets.html
    """
    # создаю книгу
    book = openpyxl.Workbook()

    # по умолчанию создается с таблицей Sheet
    # print(book.sheetnames)
    book.remove(book.active)
    # создаю таблицы
    name_sheet = "Общая информация"
    sheet = book.create_sheet(name_sheet)

    # print(data_samples)
    for row in data_samples:  # получаю данные
        # book.guess_types = True
        sheet.append(row)  # записываю данные в строки таблицы
    sheet.insert_rows(0, amount=2)
    sheet["B1"].value = name_sheet.upper()
    sheet["B1"].font = Font(bold=True)

    sheet.column_dimensions["B"].width = 70  # прим. колво символов
    sheet.column_dimensions["C"].width = 60
    set_border(sheet, "B3:C14")
    set_color(sheet, "A3:A14")
    set_width(sheet)
    # РАСХОД МАТЕРИАЛОВ, ИСПОЛЬЗУЕМЫХ ПРИ РЕМОНТЕ СКВАЖИНЫ
    sheet.insert_rows(15, amount=2)
    sheet["B16"].value = 'РАСХОД МАТЕРИАЛОВ, ИСПОЛЬЗУЕМЫХ ПРИ РЕМОНТЕ СКВАЖИНЫ'
    sheet["B16"].font = Font(bold=True)
    idx_table2 = [''] + list(range(13, 22))
    sheet.append(idx_table2)
    set_color(sheet, "B17:J17")
    table2 = list(map(lambda x: [''] + x, table2))
    for row in table2:  # получаю данные
        # book.guess_types = True
        sheet.append(row)
    end_table2 = 15 + 3 + len(table2) - 1
    set_border(sheet, f"B18:J{end_table2}")
    # ПЕРФОРАЦИЯ, ОТКЛЮЧЕНИЕ ПЛАСТОВ
    start_table3 = 15 + 3 + len(table2)
    sheet.insert_rows(start_table3, amount=2)
    sheet[f"B{start_table3 + 1}"].value = 'ПЕРФОРАЦИЯ, ОТКЛЮЧЕНИЕ ПЛАСТОВ'
    sheet[f"B{start_table3 + 1}"].font = Font(bold=True)
    idx_table3 = [''] + list(range(22, 33))
    sheet.append(idx_table3)
    set_color(sheet, f"B{start_table3 + 2}:L{start_table3 + 2}")
    table3 = list(map(lambda x: [''] + x, table3))
    for row in table3:  # получаю данные
        # book.guess_types = True
        sheet.append(row)
    set_border(sheet, f"B{start_table3 + 3}:L{start_table3 + 3 + len(table3) - 1}")
    sheet.merge_cells(f"G{start_table3 + 3}:H{start_table3 + 3}")
    sheet[f"G{start_table3 + 3}"].alignment = Alignment(horizontal='center')
    merge_2_row(sheet, 2, 6, start_table3 + 3)
    merge_2_row(sheet, 9, 12, start_table3 + 3)
    path_result = f'Быстрый_тест_таблицы/{path[:-3]}xlsx'
    os.makedirs(os.path.dirname(path_result), exist_ok=True)
    book.save(path_result)
    # book = openpyxl.load_workbook(path_result)
    # sheet = book.active
    # sheet = book[name_sheet]
    # col = sheet.column_dimensions['A']
    # col.font = Font(color="00FF00")
    # book.save(path_result)
    return path_result


if __name__ == '__main__':
    with open('path_directory.txt', encoding='utf-8', mode='r') as f:
        directory = f.readline()
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                path_file = os.path.join(root, file)
                data, table2, table3 = get_info_from_text(path_file)
                create_xlsx(path_file, data, table2, table3)
                break
