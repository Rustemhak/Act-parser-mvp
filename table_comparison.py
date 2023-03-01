from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill

redFill = PatternFill(start_color='ffb3b9',
                   end_color='ffb3b9',
                   fill_type='solid')
def get_header(from_algo, benchmark):
    result = []
    algo_headers = list(from_algo.values)[0]
    bench_headers = list(benchmark.values)[0]

    for i in range(len(algo_headers)):
        for j in range(len(bench_headers)):
            if str(algo_headers[i]).replace(' ', '') == str(bench_headers[j]).replace(' ', ''):
                result.append((i, j))
                break

    return result


def get_ws(path_from_algo, path_benchmark):
    table_from_algo = load_workbook(path_from_algo)['Лист1']
    table_benchmark = load_workbook(path_benchmark)['Лист1']

    headers = get_header(table_from_algo, table_benchmark)

    table_from_algo = sorted(table_from_algo, key=lambda row: str(row[0].value))
    table_benchmark = sorted(table_benchmark, key=lambda row: str(row[0].value))

    return table_from_algo, table_benchmark, headers


def clean_text(string: str) -> str:
    result = string.replace(';', '').replace(',0м3', 'м3').replace(' ', '').lower()

    if result[-1] == '%':
        try:
            result = str(float(result[:-1]) * 0.01)
        except Exception as ex:
            print(ex)

    if result.endswith('.'):
        result = result[:-1]

    return result


def set_color(ws, cell_range, boolean_matrix):
    for i, row in enumerate(ws[cell_range]):
        for j, cell in enumerate(row):
            if  not boolean_matrix[i][j]:
                cell.fill = redFill


def colored_cells(ws, boolean_matrix):
    set_color(ws, "A1:AR85", boolean_matrix)


def comparison(path_from_algo, path_benchmark, log=False):
    """
    :param path_from_algo: путь к таблице, полученной в результате работы алгоритма
    :param path_benchmark: путь к таблице, проверенной вручную
    :param log: Если True, то в консоли выводится информация по проверке: выводятся названия сравниваемых скважин,
    если названия одинаковые, то выводится знак "=" и начинается сравнение значений в ячейчках строки,
    выводятся несоответствия, если они есть в сравниваемых строках по одной скважине
    :return: возвращается матрица (список списков) True/False - маска к таблице, полученной от алгоритма,
    где на местах False несоответствия с проверенной вручную таблицей
    """
    from_algo, benchmark, headers = get_ws(path_from_algo, path_benchmark)
    result = [[True] * len(row) for row in from_algo]
    i = 0
    j = 0

    while (i < len(from_algo)) and (j < len(benchmark)):
        algo = str(from_algo[i][0].value).lower().replace('a', 'а')
        bench = str(benchmark[j][0].value).lower().replace('a', 'а')
        if log:
            print(f'from algo {algo}, benchmark {bench}')

        if algo == bench:

            if log:
                print('=')

            for p, q in headers:
                if clean_text(str(from_algo[i][p].value)) != clean_text(str(benchmark[j][q].value)):
                    # from_algo[i][p].font = Font(color='FF6666')
                    result[i][p] = False
                    if log:
                        print(f'MISTAKE from algo: {from_algo[i][p].value} from benchmark: {benchmark[j][q].value}')

            i += 1
            j += 1

        elif algo < bench:
            if log:
                print('<')
            i += 1

        elif algo > bench:
            if log:
                print('>')
            j += 1

    return result


if __name__ == '__main__':
    res = comparison('ГЭР-v1.xlsx', "таблица Гидрофобная_эмульсия.xlsx", log=True)
    for row in res:
        print(row)
    wb = load_workbook('ГЭР-v1.xlsx')
    table_from_algo = wb['Лист1']
    colored_cells(table_from_algo, res)
    wb.save('ГЭР-v1.xlsx')
    # print(f'{clean_text("0.005%")} == {clean_text("5e-05")} is {clean_text("0.005%") == clean_text("5e-05")}')
